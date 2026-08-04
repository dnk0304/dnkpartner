#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Regenerate the INE municipality gazetteer from the OFFICIAL INE register.

    python scripts/build-ine-gazetteer.py            # fetch + rebuild
    python scripts/build-ine-gazetteer.py --year 26  # pin an edition

Writes:
    scraper/config/ine_municipalities.csv              (one row per municipality)
    scraper/config/ine_municipalities_coofficial.csv   (every other official form)

============================== WHY THIS EXISTS ==============================

The previous gazetteer was assembled from Wikidata `es` labels. It carried
1980s-era Castilian-only names ("Alegria de Alava" for the municipality that
has been "Alegria-Dulantzi" since 1989; "Aramayona" for "Aramaio"), invented
sub-municipal rows with 11-digit codes, and no provenance. Everything on the
site resolves against this file, so an unofficial, undated gazetteer is worse
than a merely old one -- it gets trusted.

THE SOURCE IS THE AUTHORITY. This script fetches INE's own municipality
dictionary and derives everything from it. It does not merge in Wikipedia, a
scrape, or a third-party aggregate. If a municipality is wrong here, it is
wrong at INE, and that is a fact we can cite rather than a mystery.

=============================== THE RULES ===================================

1. OFFICIAL SOURCE ONLY.
   https://www.ine.es/daco/daco42/codmun/diccionario{YY}.xlsx  --  INE's
   "Relacion de municipios y codigos por comunidades autonomas y provincias".
   Columns: CODAUTO, CPRO, CMUN, DC, NOMBRE.

2. MUNICIPALITIES ONLY.
   This register contains municipalities and nothing else -- no pedanias, no
   EATIMs, no entidades singulares. That is precisely why it is the right
   source: the junk-entry class simply cannot enter. The 5-digit INE code
   (CPRO + CMUN) is the stable key; every emitted row carries it.

3. CO-OFFICIAL NAMES ARE DERIVED, NOT INVENTED.
   INE's NOMBRE field IS the full legal denomination. For bilingual
   municipalities it carries every official form, slash-separated:
   "Agurain/Salvatierra", "Donostia/San Sebastian", "Elx/Elche". Splitting on
   "/" yields the official forms -- each one is legally the name of that
   municipality, not an alias someone decided to add.
   INE also writes leading articles inverted: "Coruna, A", "Vila Joiosa, la".
   Un-inverting is a documented, mechanical transform over a closed article
   set, applied only when the post-comma tail IS one of those articles -- so
   real internal commas ("Cruilles, Monells i Sant Sadurni de l'Heura") are
   left alone.

4. NO HAND EDITS.
   There is no exception list in this file today. If one ever becomes
   necessary it lives HERE, in code, with the reason written next to it --
   never as a silent edit to the generated CSV.

5. PROVINCE COMES FROM THE CODE, NOT THE NAME.
   CPRO -> canonical province name via scraper/config/provinces.py. Names are
   ambiguous and change; the numeric code does not.
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import io
import os
import sys
import urllib.request

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_DIR = os.path.join(REPO, "scraper", "config")
sys.path.insert(0, CONFIG_DIR)

import provinces as _provinces  # noqa: E402

SOURCE_TEMPLATE = "https://www.ine.es/daco/daco42/codmun/diccionario{yy}.xlsx"
SOURCE_PAGE = "https://www.ine.es/dyngs/INEbase/es/operacion.htm?c=Estadistica_C&cid=1254736177031"
UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)

# The closed set of leading articles INE writes inverted after a comma. A tail
# outside this set means the comma is part of the municipality's real name.
# Keyed lowercase, accent-free; the emitted variant preserves INE's own casing.
TRAILING_ARTICLES = {
    "el", "la", "los", "las", "l'", "els", "les",
    "a", "as", "o", "os", "lo", "es", "sa", "ses", "sos",
}


def fetch(url: str) -> tuple[bytes, str]:
    """Download the register. Returns (bytes, Last-Modified header)."""
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as resp:
        if resp.status != 200:
            raise SystemExit(f"INE returned HTTP {resp.status} for {url}")
        return resp.read(), resp.headers.get("Last-Modified", "")


def uninvert(name: str) -> str | None:
    """"Coruna, A" -> "A Coruna"; "Orxa, l'" -> "l'Orxa". None if not inverted."""
    comma = name.rfind(",")
    if comma <= 0:
        return None
    head, tail = name[:comma].strip(), name[comma + 1:].strip()
    key = tail.lower().replace("’", "'")
    if not head or key not in TRAILING_ARTICLES:
        return None
    return f"{tail}{head}" if key.endswith("'") else f"{tail} {head}"


def official_forms(nombre: str) -> list[str]:
    """Every official written form of a municipality's name, in INE's order.

    The first element is the primary denomination (what we display); the rest
    are equally official and must all be matchable.
    """
    forms: list[str] = []
    for part in nombre.split("/"):
        part = part.strip()
        if not part:
            continue
        # Only the un-inverted form is emitted. "Coruna, A" is not a second
        # name, it is INE's alphabetical FILING convention for the one name
        # "A Coruna" -- and emitting both would register two different
        # spellings under one normalized key, which the resolver reads as
        # "this key is ambiguous" and then refuses to name the town at all.
        # The raw INE string is preserved in the `nombre_ine` column, and both
        # matchers (municipalityKey, _word_variants) already un-invert input.
        forms.append(uninvert(part) or part)
    out: list[str] = []
    for f in forms:
        if f not in out:
            out.append(f)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", default=None,
                    help="2-digit INE edition (default: try current year, fall back)")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        raise SystemExit("pip install openpyxl")

    candidates = ([args.year] if args.year
                  else [f"{y % 100:02d}" for y in range(_dt.date.today().year, 2019, -1)])
    blob = url = last_modified = None
    for yy in candidates:
        u = SOURCE_TEMPLATE.format(yy=yy)
        try:
            blob, last_modified = fetch(u)
            url = u
            break
        except Exception:
            continue
    if blob is None:
        raise SystemExit("could not fetch any INE municipality dictionary edition")

    digest = hashlib.sha256(blob).hexdigest()
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Row 0 is the register's own title -- it states the reference date
    # ("...a 1 de enero de 2026"). Carry it verbatim; it is the answer to
    # "how old is this file?".
    title = str(rows[0][0]).strip()
    header = [str(c).strip() if c else "" for c in rows[1]]
    if header[:5] != ["CODAUTO", "CPRO", "CMUN", "DC", "NOMBRE"]:
        raise SystemExit(f"unexpected INE columns: {header!r} -- inspect before trusting")

    records = []
    for r in rows[2:]:
        if not r or not r[1] or not r[2] or not r[4]:
            continue
        cpro, cmun, nombre = str(r[1]).strip(), str(r[2]).strip(), str(r[4]).strip()
        code = f"{cpro.zfill(2)}{cmun.zfill(3)}"
        if len(code) != 5 or not code.isdigit():
            raise SystemExit(f"malformed INE code from row {r!r}")
        prov = _provinces.province_by_code_strict(cpro.zfill(2))
        if not prov:
            raise SystemExit(f"no canonical province for CPRO {cpro!r} ({nombre})")
        records.append((code, nombre, prov))

    if len(records) < 8000:
        raise SystemExit(f"only {len(records)} municipalities parsed -- refusing to write")
    if len({c for c, _, _ in records}) != len(records):
        raise SystemExit("duplicate INE codes in the register -- refusing to write")

    stamp = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    banner = [
        "# SOURCE: Instituto Nacional de Estadistica (INE), official register.",
        f"# TITLE:  {title}",
        f"# URL:    {url}",
        f"# INE Last-Modified: {last_modified}",
        f"# sha256: {digest}",
        f"# GENERATED: {stamp} by scripts/build-ine-gazetteer.py -- DO NOT HAND-EDIT.",
        f"# LANDING PAGE: {SOURCE_PAGE}",
    ]

    primary_path = os.path.join(CONFIG_DIR, "ine_municipalities.csv")
    coof_path = os.path.join(CONFIG_DIR, "ine_municipalities_coofficial.csv")

    n_primary = n_coof = 0
    with open(primary_path, "w", encoding="utf-8", newline="") as fh:
        fh.write("\n".join(banner + [
            "# CONTENT: one row per municipality -- the primary official denomination.",
            "#          Municipalities ONLY; this register contains no pedanias or",
            "#          sub-municipal entities. `ine` is the 5-digit CPRO+CMUN code.",
            "",
        ]) + "\n")
        w = csv.writer(fh)
        w.writerow(["ine", "municipio", "provincia", "nombre_ine"])
        for code, nombre, prov in records:
            w.writerow([code, official_forms(nombre)[0], prov, nombre])
            n_primary += 1

    with open(coof_path, "w", encoding="utf-8", newline="") as fh:
        fh.write("\n".join(banner + [
            "# CONTENT: the ADDITIONAL official written forms of a municipality's name",
            "#          (the co-official-language denominations INE stores slash-separated).",
            "#          Every row is derived mechanically from the `nombre_ine` of the",
            "#          same INE code in ine_municipalities.csv -- nothing is invented.",
            "#          `forma`: coofficial = a slash-separated official denomination.",
            "",
            "",
        ]) + "\n")
        w = csv.writer(fh)
        w.writerow(["ine", "municipio", "provincia", "forma", "nombre_ine"])
        for code, nombre, prov in records:
            forms = official_forms(nombre)
            for f in forms[1:]:
                w.writerow([code, f, prov,
                            "coofficial", nombre])
                n_coof += 1

    print(f"source     : {url}")
    print(f"title      : {title}")
    print(f"sha256     : {digest}")
    print(f"municipios : {n_primary}  -> {os.path.relpath(primary_path, REPO)}")
    print(f"alt forms  : {n_coof}  -> {os.path.relpath(coof_path, REPO)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
